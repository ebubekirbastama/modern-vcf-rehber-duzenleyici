# ================= AUTO DEPENDENCY INSTALL =================
import sys
import subprocess

REQUIRED_PACKAGES = [
    "ttkbootstrap==1.14.6",
    "Pillow",
    "openpyxl"
]

def install_and_import(package):
    try:
        __import__(package.split("==")[0])
    except ImportError:
        print(f"[+] {package} bulunamadı, yükleniyor...")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", package],
            stdout=subprocess.DEVNULL
        )

for pkg in REQUIRED_PACKAGES:
    install_and_import(pkg)
import tkinter as tk
from tkinter import filedialog, messagebox, Toplevel
from ttkbootstrap import Style
from ttkbootstrap.widgets import *
import re
import quopri
from tkinter import StringVar
from openpyxl import Workbook, load_workbook
from datetime import datetime

contacts = []
filtered_contacts = []
stats = {
    "total_contacts": 0,
    "duplicates_removed": 0,
    "original_count": 0
}


# ----------------- GÜÇLÜ VCF IMPORTER -----------------
def decode_value(value: str):
    """QUOTED-PRINTABLE ve UTF-8 decoding"""
    try:
        if "=" in value:
            value = quopri.decodestring(value).decode("utf-8")
    except:
        pass
    return value.strip()


def normalize_phone(num: str):
    """Bütün telefon formatlarını +90 555 555 55 55 olacak şekilde normalize eder"""
    digits = re.sub(r"[^0-9+]", "", num)

    # Başındaki 0'yı kırp
    if digits.startswith("0"):
        digits = digits[1:]

    # Eğer + yoksa → Türkiye varsay
    if not digits.startswith("+"):
        if digits.startswith("90"):
            digits = "+" + digits
        elif len(digits) == 10:  # Örn: 5323456789
            digits = "+90" + digits
        else:
            digits = "+" + digits

    # Formatlı çıkış
    d = re.sub(r"[^0-9]", "", digits)
    if d.startswith("90") and len(d) == 12:
        return f"+90 {d[2:5]} {d[5:8]} {d[8:10]} {d[10:12]}"

    return digits


def read_vcf(file_path):
    """Senin GUI koduna tam uyumlu yeni profesyonel VCF importer"""
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()

    cards = re.findall(r"BEGIN:VCARD(.*?)END:VCARD", content, re.S)

    unique = []
    numbers_seen = set()
    duplicate_count = 0

    for c in cards:
        name = ""
        org = ""
        numbers = []

        for line in c.split("\n"):
            line = line.strip()

            # FN
            if line.startswith("FN"):
                parts = line.split(":", 1)
                if len(parts) > 1:
                    name = decode_value(parts[1])

            # N alanı fallback
            elif line.startswith("N:") or line.startswith("N;"):
                parts = line.split(":", 1)
                if len(parts) > 1:
                    clean = decode_value(parts[1].replace(";", " ").strip())
                    if not name and clean:
                        name = clean

            # ORG fallback
            elif line.startswith("ORG"):
                parts = line.split(":", 1)
                if len(parts) > 1:
                    org = decode_value(parts[1])

            # TEL alanı
            elif line.startswith("TEL"):
                parts = line.split(":", 1)
                if len(parts) > 1:
                    raw = parts[1].strip()
                    normalized = normalize_phone(raw)
                    numbers.append(normalized)

        # İsim yok → fallback üret
        if not name:
            if org:
                name = org
            elif numbers:
                name = f"Kişi {numbers[0]}"
            else:
                name = "İsimsiz"

        # Aynı numarayı tekrar eklememek
        for num in numbers:
            if num in numbers_seen:
                duplicate_count += 1
                continue
            numbers_seen.add(num)

            unique.append({"name": name, "number": num})

    return unique, duplicate_count


# ----------------- TABLOYU YENİLE -----------------
def refresh_table(data=None):
    global filtered_contacts
    tree.delete(*tree.get_children())

    if data is None:
        data = contacts

    filtered_contacts = data

    for idx, c in enumerate(data):
        tree.insert("", "end", iid=str(idx), values=(c["name"], c["number"]))


# ----------------- ARAMA -----------------
def search_contacts(event=None):
    text = search_var.get().lower()

    if text == "":
        refresh_table(contacts)
        return

    results = [c for c in contacts if text in c["name"].lower() or text in c["number"]]
    refresh_table(results)


# ----------------- DÜZENLE -----------------
def edit_contact():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Uyarı", "Düzenlemek için bir kayıt seç!")
        return

    idx = int(selected[0])
    contact = filtered_contacts[idx]

    edit_win = Toplevel(root)
    edit_win.title("Kişiyi Düzenle")
    edit_win.geometry("300x200")

    Label(edit_win, text="Ad").pack(pady=5)
    name_entry = Entry(edit_win)
    name_entry.insert(0, contact["name"])
    name_entry.pack()

    Label(edit_win, text="Numara").pack(pady=5)
    number_entry = Entry(edit_win)
    number_entry.insert(0, contact["number"])
    number_entry.pack()

    def save_edit():
        new_name = name_entry.get().strip()
        new_number = number_entry.get().strip()

        if new_name == "" or new_number == "":
            messagebox.showerror("Hata", "Alanlar boş olamaz!")
            return

        original_index = contacts.index(contact)
        contacts[original_index] = {"name": new_name, "number": new_number}

        search_contacts()
        edit_win.destroy()

    Button(edit_win, text="Kaydet", command=save_edit).pack(pady=15)


# ----------------- SİL -----------------
def delete_contact():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Uyarı", "Silmek için bir kayıt seç!")
        return

    idx = int(selected[0])
    contact = filtered_contacts[idx]

    if messagebox.askyesno("Sil", f"{contact['name']} silinsin mi?"):
        contacts.remove(contact)
        search_contacts()


# ----------------- VCF YÜKLE -----------------
def load_vcf():
    global contacts, stats
    file_path = filedialog.askopenfilename(filetypes=[("VCF Files", "*.vcf")])
    if not file_path:
        return

    contacts, dup = read_vcf(file_path)
    
    # İstatistikleri güncelle
    stats["total_contacts"] = len(contacts)
    stats["duplicates_removed"] = dup
    stats["original_count"] = len(contacts) + dup

    refresh_table()

    messagebox.showinfo("Tamamlandı", f"{dup} adet mükerrer numara kaldırıldı.")


# ----------------- İSTATİSTİKLER -----------------
def show_statistics():
    if not contacts:
        messagebox.showwarning("Uyarı", "Önce bir VCF dosyası yükleyin!")
        return
    
    # İstatistikleri hesapla
    total = len(contacts)
    duplicates = stats.get("duplicates_removed", 0)
    original = stats.get("original_count", total)
    
    # İsim istatistikleri
    names = [c["name"] for c in contacts]
    longest_name = max(names, key=len) if names else ""
    shortest_name = min(names, key=len) if names else ""
    avg_name_length = sum(len(n) for n in names) / len(names) if names else 0
    
    # Numara formatı analizi
    turkish_mobile = sum(1 for c in contacts if c["number"].startswith("+90 5"))
    other_numbers = total - turkish_mobile
    
    # İsimsiz kişiler
    unnamed = sum(1 for c in contacts if c["name"].startswith("İsimsiz") or c["name"].startswith("Kişi +"))
    
    # İstatistik penceresi - Modern ve şık tasarım
    stats_win = Toplevel(root)
    stats_win.title("📊 Rehber İstatistikleri")
    stats_win.geometry("850x950")
    stats_win.resizable(True, True)
    stats_win.configure(bg="#1a1a1a")
    
    # Üst başlık bölümü - Gradient efekti için
    header_frame = Frame(stats_win, bootstyle="dark")
    header_frame.pack(fill="x", pady=0)
    
    title_label = Label(
        header_frame, 
        text="📊 Rehber Analizi", 
        font=("Segoe UI", 24, "bold"),
        bootstyle="info",
        foreground="#00d4ff"
    )
    title_label.pack(pady=20)
    
    subtitle_label = Label(
        header_frame,
        text="Detaylı İstatistikler ve Analiz",
        font=("Segoe UI", 11),
        bootstyle="secondary"
    )
    subtitle_label.pack(pady=(0, 15))
    
    # Ana container - Scrollable
    main_container = Frame(stats_win, bootstyle="dark")
    main_container.pack(fill="both", expand=True, padx=20, pady=10)
    
    # === KART 1: Genel İstatistikler ===
    def create_stat_card(parent, title, stats_list, icon="📊"):
        card = Frame(parent, bootstyle="secondary", relief="raised", borderwidth=2)
        card.pack(fill="x", pady=10, padx=5)
        
        # Kart başlığı
        card_header = Frame(card, bootstyle="dark")
        card_header.pack(fill="x", padx=15, pady=12)
        
        Label(
            card_header,
            text=f"{icon} {title}",
            font=("Segoe UI", 14, "bold"),
            bootstyle="light"
        ).pack(anchor="w")
        
        # Ayırıcı çizgi
        separator = Frame(card, bootstyle="info", height=2)
        separator.pack(fill="x", padx=15, pady=(0, 10))
        
        # İstatistik satırları
        for stat in stats_list:
            stat_frame = Frame(card, bootstyle="secondary")
            stat_frame.pack(fill="x", padx=15, pady=8)
            
            # Sol taraf - Label
            left_label = Label(
                stat_frame,
                text=stat["label"],
                font=("Segoe UI", 11),
                bootstyle="light"
            )
            left_label.pack(side="left")
            
            # Sağ taraf - Value
            value_style = stat.get("style", "light")
            right_label = Label(
                stat_frame,
                text=stat["value"],
                font=("Segoe UI", 12, "bold"),
                bootstyle=value_style
            )
            right_label.pack(side="right")
        
        # Alt padding
        Frame(card, height=10).pack()
    
    # Genel Bilgiler Kartı
    general_stats = [
        {"label": "📁 Orijinal Kayıt", "value": f"{original:,}", "style": "info"},
        {"label": "🗑️ Mükerrer Temizlenen", "value": f"{duplicates:,}", "style": "danger"},
        {"label": "✅ Temiz Kayıt", "value": f"{total:,}", "style": "success"},
    ]
    create_stat_card(main_container, "Genel Bilgiler", general_stats, "📋")
    
    # İsim Analizi Kartı
    name_stats = [
        {"label": "👤 En Uzun İsim", "value": f"{longest_name[:25]}..." if len(longest_name) > 25 else longest_name, "style": "warning"},
        {"label": "👤 En Kısa İsim", "value": shortest_name, "style": "warning"},
        {"label": "📏 Ortalama Uzunluk", "value": f"{avg_name_length:.1f} karakter", "style": "info"},
    ]
    create_stat_card(main_container, "İsim Analizi", name_stats, "✍️")
    
    # Numara Dağılımı Kartı
    turkish_percent = (turkish_mobile/total*100) if total > 0 else 0
    other_percent = (other_numbers/total*100) if total > 0 else 0
    
    number_stats = [
        {"label": "📱 Türk Mobil", "value": f"{turkish_mobile:,} ({turkish_percent:.1f}%)", "style": "success"},
        {"label": "🌍 Diğer Numaralar", "value": f"{other_numbers:,} ({other_percent:.1f}%)", "style": "info"},
        {"label": "❓ İsimsiz Kişi", "value": f"{unnamed:,}", "style": "warning"},
    ]
    create_stat_card(main_container, "Numara Dağılımı", number_stats, "📞")
    
    # Alt buton bölümü
    button_frame = Frame(stats_win, bootstyle="dark")
    button_frame.pack(fill="x", pady=15)
    
    Button(
        button_frame,
        text="✖ Kapat",
        command=stats_win.destroy,
        bootstyle="danger-outline",
        width=25
    ).pack(pady=10)


# ----------------- VCF KAYDET -----------------
def save_vcf():
    file_path = filedialog.asksaveasfilename(defaultextension=".vcf")
    if not file_path:
        return

    with open(file_path, "w", encoding="utf-8") as f:
        for c in contacts:
            f.write("BEGIN:VCARD\n")
            f.write("VERSION:3.0\n")
            f.write(f"FN:{c['name']}\n")
            f.write(f"TEL:{c['number']}\n")
            f.write("END:VCARD\n\n")

    messagebox.showinfo("Kayıt", "Temiz rehber kaydedildi.")


# ----------------- EXCEL İÇE AKTAR -----------------
def import_excel():
    global contacts, stats
    file_path = filedialog.askopenfilename(
        title="Excel Dosyası Seç",
        filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
    )
    if not file_path:
        return
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        imported_contacts = []
        numbers_seen = set()
        duplicate_count = 0
        
        # İlk satır başlık olduğu için 2'den başla
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            
            name = str(row[0]).strip() if row[0] else "İsimsiz"
            number = str(row[1]).strip() if row[1] else ""
            
            if not number:
                continue
            
            # Numarayı normalize et
            normalized = normalize_phone(number)
            
            # Mükerrer kontrolü
            if normalized in numbers_seen:
                duplicate_count += 1
                continue
            
            numbers_seen.add(normalized)
            imported_contacts.append({"name": name, "number": normalized})
        
        if not imported_contacts:
            messagebox.showwarning("Uyarı", "Excel dosyasında geçerli veri bulunamadı!")
            return
        
        # Mevcut rehbere ekle veya değiştir
        response = messagebox.askyesno(
            "İçe Aktarma",
            f"{len(imported_contacts)} kişi bulundu.\n\n"
            f"Mevcut rehberi değiştir mi?\n"
            f"(Hayır = Mevcut rehbere ekle)"
        )
        
        if response:  # Evet - Değiştir
            contacts = imported_contacts
            stats["original_count"] = len(imported_contacts) + duplicate_count
        else:  # Hayır - Ekle
            contacts.extend(imported_contacts)
            stats["original_count"] = stats.get("original_count", 0) + len(imported_contacts) + duplicate_count
        
        stats["total_contacts"] = len(contacts)
        stats["duplicates_removed"] = stats.get("duplicates_removed", 0) + duplicate_count
        
        refresh_table()
        messagebox.showinfo(
            "Başarılı",
            f"✅ {len(imported_contacts)} kişi içe aktarıldı\n"
            f"🗑️ {duplicate_count} mükerrer atlandı"
        )
        
    except Exception as e:
        messagebox.showerror("Hata", f"Excel dosyası okunamadı:\n{str(e)}")


# ----------------- EXCEL DIŞA AKTAR -----------------
def export_excel():
    if not contacts:
        messagebox.showwarning("Uyarı", "Dışa aktarılacak kişi yok!")
        return
    
    file_path = filedialog.asksaveasfilename(
        title="Excel Olarak Kaydet",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if not file_path:
        return
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Rehber"
        
        # Başlık satırı - Şık formatla
        headers = ["Ad Soyad", "Telefon Numarası", "Durum"]
        ws.append(headers)
        
        # Başlık stilini ayarla
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Veri satırları
        for contact in contacts:
            status = "✓ Türk Mobil" if contact["number"].startswith("+90 5") else "○ Diğer"
            ws.append([contact["name"], contact["number"], status])
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        # Tüm hücreleri ortala
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # İstatistik sayfası ekle
        stats_ws = wb.create_sheet("İstatistikler")
        stats_ws.append(["Rehber İstatistikleri"])
        stats_ws.append([])
        stats_ws.append(["Toplam Kişi", len(contacts)])
        stats_ws.append(["Mükerrer Temizlenen", stats.get("duplicates_removed", 0)])
        stats_ws.append(["Orijinal Kayıt", stats.get("original_count", len(contacts))])
        stats_ws.append([])
        stats_ws.append(["Dışa Aktarma Tarihi", datetime.now().strftime("%d.%m.%Y %H:%M")])
        
        # İstatistik sayfası formatı
        stats_ws['A1'].font = Font(bold=True, size=14)
        stats_ws.column_dimensions['A'].width = 25
        stats_ws.column_dimensions['B'].width = 20
        
        wb.save(file_path)
        messagebox.showinfo(
            "Başarılı",
            f"✅ Excel dosyası kaydedildi!\n\n"
            f"📊 {len(contacts)} kişi\n"
            f"📁 {file_path}"
        )
        
    except Exception as e:
        messagebox.showerror("Hata", f"Excel dosyası kaydedilemedi:\n{str(e)}")


# ----------------- GUI TASARIM -----------------

style = Style(theme="darkly")

root = style.master
root.title("Modern VCF Rehber Düzenleyici")
root.geometry("1080x500")

top_frame = Frame(root)
top_frame.pack(fill="x", padx=10, pady=10)

# Sol taraf - Dosya işlemleri
Button(top_frame, text="📂 VCF Yükle", command=load_vcf, bootstyle="info").pack(side="left", padx=3)
Button(top_frame, text="💾 VCF Kaydet", command=save_vcf, bootstyle="success").pack(side="left", padx=3)

# Separator
Label(top_frame, text="|", bootstyle="secondary").pack(side="left", padx=8)

# Excel işlemleri
Button(top_frame, text="📥 Excel İçe Aktar", command=import_excel, bootstyle="info-outline").pack(side="left", padx=3)
Button(top_frame, text="📤 Excel Dışa Aktar", command=export_excel, bootstyle="success-outline").pack(side="left", padx=3)

# Separator
Label(top_frame, text="|", bootstyle="secondary").pack(side="left", padx=8)

# Düzenleme işlemleri
Button(top_frame, text="✏️ Düzenle", command=edit_contact, bootstyle="warning").pack(side="left", padx=3)
Button(top_frame, text="🗑️ Sil", command=delete_contact, bootstyle="danger").pack(side="left", padx=3)
Button(top_frame, text="📊 İstatistikler", command=show_statistics, bootstyle="primary").pack(side="left", padx=3)

# Sağ taraf - Arama
search_var = StringVar()
search_entry = Entry(top_frame, textvariable=search_var, width=30)
search_entry.pack(side="right", padx=5)
search_entry.bind("<KeyRelease>", search_contacts)

Label(top_frame, text="🔍 Ara:", bootstyle="inverse").pack(side="right")

columns = ("Ad", "Numara")
tree = Treeview(root, columns=columns, show="headings", bootstyle="dark")
tree.heading("Ad", text="Ad")
tree.heading("Numara", text="Numara")
tree.pack(fill="both", expand=True, padx=10, pady=10)

root.mainloop()
